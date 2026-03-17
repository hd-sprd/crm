"""
Salesforce data import endpoint.
Supports CSV and XLSX files for: accounts, contacts, leads, opportunities (→ deals).
"""
import io
import csv
import json
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models.user import User
from app.models.account import Account, AccountType, AccountStatus
from app.models.contact import Contact
from app.models.lead import Lead, LeadSource, LeadStatus
from app.models.deal import Deal
from app.models.settings import CustomFieldDef
from app.services.auth_service import get_current_user

router = APIRouter(prefix="/import", tags=["import"])

CHUNK_SIZE = 250

ENTITY_SINGULAR = {"accounts": "account", "contacts": "contact", "leads": "lead", "deals": "deal"}

# Available DB fields per entity (used by the mapping UI)
FIELD_DEFS: dict[str, list[dict]] = {
    "accounts": [
        {"key": "name", "label": "Account Name", "required": True},
        {"key": "industry", "label": "Industry", "required": False},
        {"key": "website", "label": "Website", "required": False},
        {"key": "country", "label": "Country", "required": False},
        {"key": "address", "label": "Address", "required": False},
        {"key": "region", "label": "Region", "required": False},
        {"key": "segment", "label": "Segment", "required": False},
        {"key": "notes", "label": "Notes", "required": False},
    ],
    "contacts": [
        {"key": "last_name", "label": "Last Name", "required": True},
        {"key": "first_name", "label": "First Name", "required": False},
        {"key": "email", "label": "Email", "required": False},
        {"key": "phone", "label": "Phone", "required": False},
        {"key": "title", "label": "Job Title", "required": False},
        {"key": "account_name", "label": "Account Name (for linking)", "required": True},
    ],
    "leads": [
        {"key": "company_name", "label": "Company", "required": False},
        {"key": "first_name", "label": "First Name", "required": False},
        {"key": "last_name", "label": "Last Name", "required": False},
        {"key": "contact_email", "label": "Email", "required": False},
        {"key": "source", "label": "Lead Source", "required": False},
        {"key": "estimated_volume", "label": "Estimated Volume", "required": False},
        {"key": "qualification_notes", "label": "Notes", "required": False},
        {"key": "timeline", "label": "Timeline", "required": False},
    ],
    "deals": [
        {"key": "title", "label": "Deal Title", "required": True},
        {"key": "account_name", "label": "Account Name (for linking)", "required": True},
        {"key": "value_eur", "label": "Value (EUR)", "required": False},
        {"key": "probability", "label": "Probability (%)", "required": False},
        {"key": "expected_close_date", "label": "Close Date", "required": False},
        {"key": "branding_requirements", "label": "Notes / Description", "required": False},
    ],
}

# Default Salesforce column header → DB field key mapping
SF_DEFAULT_MAPPING: dict[str, dict[str, str]] = {
    "accounts": {
        "Name": "name",
        "Industry": "industry",
        "Website": "website",
        "BillingCountry": "country",
        "BillingStreet": "address",
        "Region__c": "region",
        "Segment__c": "segment",
        "Description": "notes",
    },
    "contacts": {
        "FirstName": "first_name",
        "LastName": "last_name",
        "Email": "email",
        "Phone": "phone",
        "Title": "title",
        "AccountName": "account_name",
        "Account Name": "account_name",
    },
    "leads": {
        "Company": "company_name",
        "FirstName": "first_name",
        "LastName": "last_name",
        "Email": "contact_email",
        "LeadSource": "source",
        "NumberOfEmployees": "estimated_volume",
        "Description": "qualification_notes",
        "Timeline__c": "timeline",
    },
    "deals": {
        "Name": "title",
        "AccountName": "account_name",
        "Account Name": "account_name",
        "Amount": "value_eur",
        "Probability": "probability",
        "CloseDate": "expected_close_date",
        "Description": "branding_requirements",
    },
}

LEAD_SOURCE_MAP = {
    "Web": "website",
    "Email": "email",
    "Event": "event",
    "Referral": "referral",
    "Other": "manual",
    "Internal": "manual",
    "": "manual",
}


def _parse_rows(content: bytes, filename: str) -> list[dict]:
    """Parse CSV or XLSX file content into list of dicts."""
    fname = (filename or "").lower()

    if fname.endswith(".xlsx") or fname.endswith(".xls"):
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(status_code=400, detail="openpyxl not installed; use CSV format")
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        return [
            {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
            for row in rows[1:]
        ]
    else:
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(row) for row in reader]


def _parse_preview_rows(content: bytes, filename: str) -> tuple[list[str], list[dict], int]:
    """Parse file and return (headers, sample_rows[:3], total_data_rows)."""
    fname = (filename or "").lower()

    if fname.endswith(".xlsx") or fname.endswith(".xls"):
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(status_code=400, detail="openpyxl not installed; use CSV format")
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], [], 0
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        data_rows = [
            {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
            for row in rows[1:]
        ]
        return headers, data_rows[:3], len(data_rows)
    else:
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        headers = list(reader.fieldnames or [])
        sample: list[dict] = []
        total = 0
        for row in reader:
            if total < 3:
                sample.append(dict(row))
            total += 1
        return headers, sample, total


def _extract_custom_fields(row: dict) -> dict | None:
    """Extract custom field values from a normalized row (keys prefixed with cf__)."""
    cf = {k[4:]: v for k, v in row.items() if k.startswith("cf__") and v not in (None, "")}
    return cf if cf else None


def _normalize_row(row: dict, mapping: dict) -> dict:
    """Apply field mapping {csv_header: db_field} → new row keyed by db field names."""
    result: dict = {}
    for csv_key, db_key in mapping.items():
        if db_key and csv_key in row:
            result[db_key] = row[csv_key]
    return result


async def _import_accounts(rows: list[dict], db: AsyncSession) -> tuple[int, int, list[str]]:
    imported, skipped, errors = 0, 0, []
    for i, row in enumerate(rows):
        name = row.get("name", "").strip()
        if not name:
            skipped += 1
            continue
        existing = await db.execute(select(Account).where(Account.name == name))
        if existing.scalar_one_or_none():
            skipped += 1
            continue
        try:
            acc = Account(
                name=name,
                type=AccountType.b2b,
                industry=row.get("industry") or None,
                website=row.get("website") or None,
                address=row.get("address") or None,
                country=row.get("country") or None,
                region=row.get("region") or None,
                segment=row.get("segment") or None,
                notes=row.get("notes") or None,
                status=AccountStatus.active,
                custom_fields=_extract_custom_fields(row),
            )
            db.add(acc)
            imported += 1
        except Exception as e:
            errors.append(f"Row {i+2}: {e}")
    return imported, skipped, errors


async def _import_contacts(rows: list[dict], db: AsyncSession) -> tuple[int, int, list[str]]:
    imported, skipped, errors = 0, 0, []
    for i, row in enumerate(rows):
        last_name = row.get("last_name", "").strip()
        if not last_name:
            skipped += 1
            continue
        try:
            account_id = None
            acc_name = row.get("account_name", "")
            if acc_name:
                acc_res = await db.execute(select(Account).where(Account.name == acc_name))
                acc = acc_res.scalar_one_or_none()
                if acc:
                    account_id = acc.id

            if not account_id:
                skipped += 1
                errors.append(f"Row {i+2}: No matching account '{acc_name}' – skipped")
                continue

            contact = Contact(
                account_id=account_id,
                first_name=row.get("first_name", "").strip() or "Unknown",
                last_name=last_name,
                email=row.get("email") or None,
                phone=row.get("phone") or None,
                title=row.get("title") or None,
                is_primary=False,
                custom_fields=_extract_custom_fields(row),
            )
            db.add(contact)
            imported += 1
        except Exception as e:
            errors.append(f"Row {i+2}: {e}")
    return imported, skipped, errors


async def _import_leads(rows: list[dict], db: AsyncSession) -> tuple[int, int, list[str]]:
    imported, skipped, errors = 0, 0, []
    for i, row in enumerate(rows):
        company = row.get("company_name", "").strip()
        first = row.get("first_name", "").strip()
        last = row.get("last_name", "").strip()
        if not company and not last:
            skipped += 1
            continue
        try:
            sf_source = (row.get("source") or "").strip()
            source_key = LEAD_SOURCE_MAP.get(sf_source, "manual")
            try:
                LeadSource(source_key)
            except ValueError:
                source_key = "manual"

            volume_raw = row.get("estimated_volume", "")
            try:
                volume = int(float(volume_raw)) if volume_raw else None
            except (ValueError, TypeError):
                volume = None

            lead = Lead(
                source=LeadSource(source_key),
                status=LeadStatus.new,
                company_name=company or None,
                contact_name=f"{first} {last}".strip() or None,
                contact_email=row.get("contact_email") or None,
                qualification_notes=row.get("qualification_notes") or None,
                estimated_volume=volume,
                timeline=row.get("timeline") or None,
                custom_fields=_extract_custom_fields(row),
            )
            db.add(lead)
            imported += 1
        except Exception as e:
            errors.append(f"Row {i+2}: {e}")
    return imported, skipped, errors


async def _import_deals(rows: list[dict], db: AsyncSession) -> tuple[int, int, list[str]]:
    imported, skipped, errors = 0, 0, []
    for i, row in enumerate(rows):
        title = row.get("title", "").strip()
        if not title:
            skipped += 1
            continue
        try:
            acc_name = row.get("account_name", "")
            account_id = None
            if acc_name:
                acc_res = await db.execute(select(Account).where(Account.name == acc_name.strip()))
                acc = acc_res.scalar_one_or_none()
                if acc:
                    account_id = acc.id

            if not account_id:
                skipped += 1
                errors.append(f"Row {i+2}: No matching account '{acc_name}' – skipped")
                continue

            val_raw = row.get("value_eur", "")
            try:
                value = float(str(val_raw).replace(",", ".")) if val_raw else None
            except (ValueError, TypeError):
                value = None

            prob_raw = row.get("probability", "")
            try:
                prob = int(float(prob_raw)) if prob_raw else 0
            except (ValueError, TypeError):
                prob = 0

            close_date = None
            close_raw = row.get("expected_close_date", "")
            if close_raw:
                for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d.%m.%Y"):
                    try:
                        close_date = datetime.strptime(close_raw.strip(), fmt).date()
                        break
                    except ValueError:
                        continue

            deal = Deal(
                title=title,
                account_id=account_id,
                stage="lead_received",
                value_eur=value,
                probability=prob,
                expected_close_date=close_date,
                branding_requirements=row.get("branding_requirements") or None,
                custom_fields=_extract_custom_fields(row),
            )
            db.add(deal)
            imported += 1
        except Exception as e:
            errors.append(f"Row {i+2}: {e}")
    return imported, skipped, errors


@router.post("/preview")
async def preview_import(
    object_type: str = Query(..., description="accounts | contacts | leads | deals"),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return headers, sample rows, suggested field mapping, and available DB fields."""
    if object_type not in ("accounts", "contacts", "leads", "deals"):
        raise HTTPException(status_code=400, detail="object_type must be one of: accounts, contacts, leads, deals")

    content = await file.read()
    if len(content) > 50 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (max 50 MB)")

    try:
        headers, sample_rows, total_rows = _parse_preview_rows(content, file.filename or "")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

    default_map = SF_DEFAULT_MAPPING.get(object_type, {})
    suggested_mapping = {h: default_map.get(h) for h in headers}

    # Load custom field definitions for this entity type
    entity = ENTITY_SINGULAR[object_type]
    cf_result = await db.execute(
        select(CustomFieldDef)
        .where(CustomFieldDef.applies_to == entity, CustomFieldDef.is_active == True)
        .order_by(CustomFieldDef.field_order)
    )
    custom_field_defs = cf_result.scalars().all()
    custom_available = [
        {
            "key": f"cf__{cf.name}",
            "label": f"{cf.label_en} (custom)",
            "required": cf.is_required,
        }
        for cf in custom_field_defs
    ]

    return {
        "headers": headers,
        "sample_rows": sample_rows,
        "suggested_mapping": suggested_mapping,
        "available_fields": FIELD_DEFS.get(object_type, []) + custom_available,
        "total_rows": total_rows,
    }


@router.post("/salesforce")
async def import_salesforce(
    object_type: str = Query(..., description="accounts | contacts | leads | deals"),
    file: UploadFile = File(...),
    field_mapping: str | None = Form(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if object_type not in ("accounts", "contacts", "leads", "deals"):
        raise HTTPException(status_code=400, detail="object_type must be one of: accounts, contacts, leads, deals")

    content = await file.read()
    if len(content) > 50 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (max 50 MB)")

    try:
        rows = _parse_rows(content, file.filename or "")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

    if not rows:
        return {"imported": 0, "skipped": 0, "errors": [], "message": "File is empty"}

    if field_mapping:
        try:
            mapping: dict = json.loads(field_mapping)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid field_mapping JSON")
    else:
        mapping = SF_DEFAULT_MAPPING.get(object_type, {})

    normalized_rows = [_normalize_row(row, mapping) for row in rows]

    handlers = {
        "accounts": _import_accounts,
        "contacts": _import_contacts,
        "leads": _import_leads,
        "deals": _import_deals,
    }
    handler = handlers[object_type]

    imported, skipped, errors = 0, 0, []
    for chunk_start in range(0, len(normalized_rows), CHUNK_SIZE):
        chunk = normalized_rows[chunk_start:chunk_start + CHUNK_SIZE]
        imp, skp, errs = await handler(chunk, db)
        imported += imp
        skipped += skp
        errors.extend(errs)
        await db.commit()

    return {
        "object_type": object_type,
        "total_rows": len(rows),
        "imported": imported,
        "skipped": skipped,
        "errors": errors[:50],
        "message": f"Import complete: {imported} imported, {skipped} skipped, {len(errors)} errors",
    }


@router.get("/export/{object_type}")
async def export_csv(
    object_type: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export CRM data as CSV. object_type: accounts | contacts | leads | deals"""
    if object_type not in ("accounts", "contacts", "leads", "deals"):
        raise HTTPException(status_code=400, detail="object_type must be one of: accounts, contacts, leads, deals")

    output = io.StringIO()
    writer = csv.writer(output)

    if object_type == "accounts":
        writer.writerow(["id", "name", "type", "status", "industry", "country", "region", "segment", "website", "address", "notes", "created_at"])
        result = await db.execute(select(Account).order_by(Account.name))
        for acc in result.scalars().all():
            writer.writerow([acc.id, acc.name, acc.type, acc.status, acc.industry, acc.country, acc.region, acc.segment, acc.website, acc.address, acc.notes, acc.created_at])

    elif object_type == "contacts":
        writer.writerow(["id", "first_name", "last_name", "email", "phone", "title", "account_name", "is_primary", "created_at"])
        result = await db.execute(
            select(Contact).options(selectinload(Contact.account)).order_by(Contact.last_name)
        )
        for c in result.scalars().all():
            writer.writerow([c.id, c.first_name, c.last_name, c.email, c.phone, c.title, c.account.name if c.account else "", c.is_primary, c.created_at])

    elif object_type == "leads":
        writer.writerow(["id", "company_name", "contact_name", "contact_email", "source", "status", "estimated_volume", "timeline", "qualification_notes", "created_at"])
        result = await db.execute(select(Lead).order_by(Lead.created_at.desc()))
        for lead in result.scalars().all():
            writer.writerow([lead.id, lead.company_name, lead.contact_name, lead.contact_email, lead.source, lead.status, lead.estimated_volume, lead.timeline, lead.qualification_notes, lead.created_at])

    elif object_type == "deals":
        writer.writerow(["id", "title", "account_name", "stage", "value_eur", "probability", "expected_close_date", "created_at"])
        result = await db.execute(
            select(Deal).options(selectinload(Deal.account)).order_by(Deal.created_at.desc())
        )
        for deal in result.scalars().all():
            writer.writerow([deal.id, deal.title, deal.account.name if deal.account else "", deal.stage, deal.value_eur, deal.probability, deal.expected_close_date, deal.created_at])

    csv_content = output.getvalue()
    return StreamingResponse(
        iter([csv_content]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={object_type}_export.csv"},
    )
